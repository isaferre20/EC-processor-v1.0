from flask import Flask, request, render_template, redirect, url_for, send_file
from werkzeug.utils import secure_filename
import os
import re
import pdfplumber
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Alignment, Font, Protection
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.cell_range import CellRange
from copy import copy
from threading import Timer
import webbrowser
import re

def sanitize_filename(name):
    return re.sub(r'[<>:"/\\|?*]', '-', name)

app = Flask(__name__)

# Configuration for file storage
app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'uploads')  # Temporary PDF upload path
app.config['EXCEL_FOLDER'] = os.path.join(app.root_path, 'excel_files')  # Folder to save generated Excel files
app.config['MODEL_FILE_PATH'] = os.path.join(app.root_path, 'static', 'EC MODELLO.xlsx')  # Path to model in static

# Ensure folders exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['EXCEL_FOLDER'], exist_ok=True)


@app.route('/')
def index():
    # Render the main page and check if there is an error to display
    error = request.args.get('error', '')
    return render_template('index.html', error=error)

@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files['file']
    base_path = os.path.abspath(request.form['base_path'])
    costo_orario_ditta = float(request.form['costo_orario'])

    # Check if a file was uploaded
    if file.filename == '':
        print("No file selected")
        return redirect(url_for('index', error="No file selected"))

    if file and file.filename.endswith('.pdf'):
        filename = secure_filename(file.filename)
        
        # Extract data from the PDF without saving in UPLOAD_FOLDER
        data, error_message = extract_pdf_data(file)
        
        if error_message:
            print(f"Error extracting data from PDF: {error_message}")
            return redirect(url_for('index', error=error_message))

        # Process PDF and generate Excel
        excel_path, file_renamed_or_old_format = process_pdf(file, filename, costo_orario_ditta, base_path=base_path, data=data)

        # Check if the Excel file is None due to an old format
        if excel_path is None and file_renamed_or_old_format:
            print("⚠ Incompatible Excel format detected.")
            return render_template('upload_success.html', 
                                   folder_path=None, 
                                   base_path=base_path, 
                                   data=data, 
                                   file_renamed=False, 
                                   old_format=True)

        # Proceed with the normal flow if no format issue
        folder_path = os.path.dirname(excel_path)
        return render_template('upload_success.html', 
                               folder_path=folder_path, 
                               base_path=base_path, 
                               data=data, 
                               file_renamed=file_renamed_or_old_format)

    else:
        print("Invalid file type. Please upload a PDF file.")
        return redirect(url_for('index', error="Invalid file type. Please upload a PDF file."))
def extract_pdf_data(file):
    """Extracts structured data from the PDF using regex patterns."""
    try:
        with pdfplumber.open(file) as pdf:
            text = " ".join(page.extract_text() for page in pdf.pages)
        text = re.sub(r'\s+', ' ', text)
        
        # Define patterns for extracting data
        patterns = {
            "client_name": r"Data Inizio Lavori:\s+\d{2}/\d{2}/\d{2}\s+([\w\sàèéìòùçÀÈÉÌÒÙÇA-Za-z/w/\.]+)\s+Imponibile",
            "date": r"Data Inizio Lavori:\s+(\d{2}/\d{2}/\d{2})",
            "scheda_num": r"Scheda num:\s+(\d+)",
            "valore_tot_manodopera": r"Valore Tot\. manodopera\s+€\s*([\d,.]+)",
            "valore_tot_materiali": r"Valore Tot\. materiali\s+€\s*([\d,.]+)",
            "costo_manodopera": r"Costo Tot\. manodopera\s+€\s*([\d,.]+)",
            "costo_materiale": r"Costo Tot\. Materiali:\s+€\s*([\d,.]+)"
        }

        data = {key: (re.search(pattern, text).group(1).strip() if re.search(pattern, text) else None)
                for key, pattern in patterns.items()}
        
        # *Sanitize the extracted client name*
        if data.get("client_name"):
            data["client_name"] = sanitize_filename(data["client_name"])

        # Check if essential data is missing
        required_fields = ["client_name", "date", "scheda_num"]
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            print(f"Missing required fields: {missing_fields}")
            # Return an error message indicating which fields are missing
            return None, f"Error: Missing required data - {', '.join(missing_fields)}"

        # If all required data is present, return data and no error message
        return data, None
    except Exception as e:
        print(f"Exception during PDF processing: {e}")
        return None, f"Error processing PDF: {e}"

def process_pdf(file, filename, costo_orario_ditta, base_path, data):
    if not data:
        print("Data is missing, skipping process_pdf")
        return None, False

    client_name = sanitize_filename(data['client_name'])
    print(client_name)
    client_folder_name = f"EC {client_name}"
    client_folder_path = os.path.join(base_path, client_folder_name)
    calcoli_folder_path = os.path.join(client_folder_path, "CALCOLI")
    schede_costi_folder_path = os.path.join(client_folder_path, "CALCOLI/SCHEDE COSTI")
    schede_cliente_folder_path = os.path.join(client_folder_path, "CALCOLI/SCHEDE CLIENTE")
    inviati_folder_path = os.path.join(client_folder_path, "INVIATI")

    # Ensure folders exist
    os.makedirs(schede_cliente_folder_path, exist_ok=True)
    os.makedirs(schede_costi_folder_path, exist_ok=True)
    os.makedirs(inviati_folder_path, exist_ok=True)

    # Save the uploaded file to the "Schede Costi" folder
    file_renamed = False
    pdf_path = os.path.abspath(os.path.join(schede_costi_folder_path, filename))

    # Append '_new' with an incremental number if the file already exists
    base_name, ext = os.path.splitext(filename)
    counter = 1
    new_path = pdf_path

    # Increment the counter until a unique filename is found
    while os.path.exists(new_path):
        new_filename = f"{base_name}_new{counter}{ext}"
        new_path = os.path.abspath(os.path.join(schede_costi_folder_path, new_filename))
        file_renamed = True
        counter += 1

    pdf_path = new_path
    if counter > 1:
        print(f"File already exists. Saving as: {os.path.basename(pdf_path)}")

    # Ensure the file is readable before saving
    try:
        file.seek(0)  # Reset pointer
        file_content = file.read()
        print(f"File size: {len(file_content)} bytes")  # Ensure the file has content
        file.seek(0)  # Reset again for saving
    except Exception as e:
        raise ValueError(f"Error reading uploaded file: {e}")

    # Save to the destination and verify
    try:
        print(f"Saving file to: {pdf_path}")
        file.save(pdf_path)

        # Test the saved file
        with open(pdf_path, 'rb') as f:
            f.read()  # Try reading the file
        print(f"File saved and verified at {pdf_path}")
    except Exception as e:
        raise ValueError(f"Error saving file: {e}")

    # Verify the saved file exists and is readable
    if not os.path.exists(pdf_path):
        raise ValueError("File was not saved properly!")

    # Define the path to the Excel file and the model template
    excel_file_path = os.path.join(calcoli_folder_path, f"EC {data['client_name'].title()}.xlsx")
    model_file_path = app.config['MODEL_FILE_PATH']

    # Check if the Excel file already exists
    if not os.path.exists(excel_file_path):
        print(f"Creating Excel file from template at {model_file_path}")
        shutil.copyfile(model_file_path, excel_file_path)

    # Load the Excel workbook and select the "SCHEDE" sheet
    wb = load_workbook(excel_file_path)

    # Check if the "SCHEDE" sheet exists
    if "SCHEDE" not in wb.sheetnames:
        print(f"⚠ Error: The Excel file '{excel_file_path}' does not have a 'SCHEDE' sheet. Possibly old format.")
        return None, True  # Returning True here signals an incompatible format

    # Select the "SCHEDE" sheet
    ws = wb["SCHEDE"]

    # Check if B1, B2, and B3 are empty to determine if the sheet is empty
    if ws["B1"].value is None and ws["B2"].value is None and ws["B3"].value is None:
        start_row = 1
    else:
        last_row = ws.max_row
        start_row = ((last_row - 1) // 23 + 1) * 23 + 1

    scheda_index = (start_row - 1) // 23

    # Insert data for the new scheda
    insert_data(ws, scheda_index, data, costo_orario_ditta, schede_costi_folder_path, pdf_path)

    wb.save(excel_file_path)
    print(f"Excel file saved at {excel_file_path}")
    return excel_file_path, file_renamed

def copy_row_format_and_formulas(ws, src_row, dest_row):
    """
    Copies the format and formulas from a source row to a destination row.
    Adjusts relative formulas to work in the new destination row.
    """
    row_offset = dest_row - src_row
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dest_cell = ws.cell(row=dest_row, column=col)

        # Copy the style
        if src_cell.has_style:
            dest_cell._style = copy(src_cell._style)
        
        # Copy the formula and adjust if necessary
        if src_cell.data_type == "f":  # 'f' means the cell contains a formula
            # Offset any cell references in the formula to match the destination row
            formula = src_cell.value
            adjusted_formula = adjust_formula(formula, row_offset)
            dest_cell.value = adjusted_formula
        else:
            dest_cell.value = src_cell.value

        # Copy the number format, if it's a date or other special format
        if src_cell.is_date:
            dest_cell.number_format = src_cell.number_format

def adjust_formula(formula, row_offset):
    """
    Adjusts a formula by shifting any row numbers by a given offset.
    """
    import re
    # Regular expression to match cell references (e.g., B7, C12)
    cell_reference_pattern = r"([A-Z]+)(\d+)"
    
    def shift_row(match):
        column, row = match.groups()
        new_row = int(row) + row_offset
        return f"{column}{new_row}"
    
    # Apply row shifting to each cell reference in the formula
    adjusted_formula = re.sub(cell_reference_pattern, shift_row, formula)
    return adjusted_formula


def insert_data(ws, scheda_index, data, costo_orario_ditta, schede_costi_folder_path, pdf_path):
    start_row = 1 + scheda_index * 23
    for src_row in range(1, 24):
        copy_row_format_and_formulas(ws, src_row, start_row + (src_row - 1))
    
    ws[f"B{start_row}"] = data["scheda_num"]
    ws[f"B{start_row + 1}"] = data["date"]
    ws[f"B{start_row + 2}"] = costo_orario_ditta
    ws[f"B{start_row + 6}"] = data["costo_manodopera"]
    ws[f"C{start_row + 6}"] = data["costo_materiale"]
    ws[f"B{start_row + 12}"] = data["valore_tot_manodopera"]
    ws[f"C{start_row + 12}"] = data["valore_tot_materiali"]

@app.route('/download/<filename>')
def download_file(filename):
    file_path = os.path.join(app.config['EXCEL_FOLDER'], filename)
    
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return "Error: File not found", 404

def open_browser():
    # ✅ Prevent opening multiple tabs
    if not os.environ.get("WERKZEUG_RUN_MAIN"):  
        webbrowser.open("http://127.0.0.1:5000")

if __name__ == '__main__':
    Timer(1, open_browser).start()
    app.run(debug=False)  # ✅ Disable debug mode in .exe
